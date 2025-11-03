from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from ..config import Config
import requests
from io import BytesIO
from .transform_schedule_data import transform_schedule
from flask import make_response

def download_excel_schedule():
    url = Config().generator_pavel_view
    response = requests.get(url)
    if response.status_code == 200:
        original_data = response.json()
        transformed_data = transform_schedule(original_data)

        wb = Workbook()
        if wb.sheetnames:
            wb.remove(wb.active)

        # Константы для ширины колонок (в символах)
        COL_WIDTHS = {
            "day": 8,    # Дни недели
            "time": 15,  # Время
            "type": 10,  # Тип недели
            "subject": 60,  # Предмет
            "room": 10   # Аудитория
        }

        GROUP_HEADER_FILL = PatternFill(start_color="E6E6FA", fill_type="solid")  # Фиолетовый для групп
        DAY_HEADER_FILL = PatternFill(start_color="F0F0F0", fill_type="solid")    # Серый для дней

        time_slots = [
            "08.30-10.00",
            "10.10-11.40",
            "11.50-13.20",
            "14.05-15.35",
            "15.55-17.25",
            "17.35-19.05",
            "19.15-20.45"
        ]

        # Стили
        header_font = Font(bold=True, size=12)
        time_font = Font(bold=True)
        type_font = Font(bold=True)
        empty_type_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

        # Границы (как раньше)
        thick_border = Border(left=Side(style="thick"), right=Side(style="thick"),
                              top=Side(style="thick"), bottom=Side(style="thick"))
        medium_border = Border(left=Side(style="medium"), right=Side(style="medium"),
                               top=Side(style="thin"), bottom=Side(style="thin"))
        thin_border = Border(left=Side(style="thin"), right=Side(style="thin"),
                             top=Side(style="thin"), bottom=Side(style="thin"))

        # === helpers для адресного проставления линий ===
        thin_side   = Side(style="thin")
        medium_side = Side(style="medium")
        thick_side  = Side(style="thick")

        def with_border(cell, l=None, r=None, t=None, b=None):
            cell.border = Border(
                left   = l or cell.border.left,
                right  = r or cell.border.right,
                top    = t or cell.border.top,
                bottom = b or cell.border.bottom,
            )
        # ================================================

        # Цвета для типов недели
        type_colors = {
            "Чс": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),  # зелёный
            "Зн": PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid"),  # синий
            "Чс/Зн": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")  # жёлтый (почти не используется теперь)
        }

        for course in range(1, 7):
            ws = wb.create_sheet(title=f"{course} Курс")
            if course not in transformed_data:
                ws.merge_cells("A1:E1")
                ws["A1"] = f"Нет данных для {course} курса"
                continue

            groups = list(transformed_data[course].keys())

            def sort_group_key(group_name):
                group_num = group_name.split('-')[1]
                return int(group_num)

            groups = sorted(groups, key=sort_group_key)
            start_col = 3

            # фиксируем ширины
            ws.column_dimensions["A"].width = COL_WIDTHS["day"]
            ws.column_dimensions["B"].width = COL_WIDTHS["time"]

            for group_idx in range(len(groups)):
                col = start_col + (group_idx * 3)
                ws.column_dimensions[get_column_letter(col)].width = COL_WIDTHS["type"]
                ws.column_dimensions[get_column_letter(col + 1)].width = COL_WIDTHS["subject"]
                ws.column_dimensions[get_column_letter(col + 2)].width = COL_WIDTHS["room"]
                ws.column_dimensions[get_column_letter(col + 1)].auto_size = False

                # Заголовки групп
                for group_idx, group in enumerate(groups):
                    col = start_col + (group_idx * 3)

                    ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col + 2)
                    header = ws.cell(row=1, column=col, value=group)
                    header.font = header_font
                    header.alignment = Alignment(horizontal="center")
                    header.border = medium_border
                    header.fill = GROUP_HEADER_FILL

                    # Подзаголовки
                    subtitles = ["Тип", "Предмет", "Ауд."]
                    for i, subtitle in enumerate(subtitles):
                        cell = ws.cell(row=2, column=col + i, value=subtitle)
                        cell.font = Font(bold=True)
                        cell.alignment = Alignment(horizontal="center")
                        cell.border = thin_border

                # Данные
                current_row = 3
                week_days = ["Понедельник", "Вторник", "Среда", "Четверг", "Пятница", "Суббота"]

                for day in week_days:
                    # День недели (вертикально)
                    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row + 13, end_column=1)
                    day_cell = ws.cell(row=current_row, column=1, value=day)
                    day_cell.alignment = Alignment(text_rotation=90, vertical="center")
                    day_cell.font = Font(bold=True, size=12)
                    day_cell.border = thick_border
                    day_cell.fill = DAY_HEADER_FILL

                    TIME_SLOT_COLORS = {
                        "08.30-10.00": "D8BFD8",
                        "10.10-11.40": "FCE5B9",
                        "11.50-13.20": "C5E0B4",
                        "14.05-15.35": "FCE5B9",
                        "15.55-17.25": "D8BFD8",
                        "17.35-19.05": "FCE5B9",
                        "19.15-20.45": "D8BFD8"
                    }

                    # Временные слоты
                    for i, time_slot in enumerate(time_slots):
                        time_row = current_row + i * 2

                        ws.merge_cells(start_row=time_row, start_column=2, end_row=time_row + 1, end_column=2)
                        time_cell = ws.cell(row=time_row, column=2, value=time_slot)
                        time_cell.font = Font(bold=True)
                        time_cell.alignment = Alignment(horizontal="center", vertical="center")
                        fill_color = TIME_SLOT_COLORS.get(time_slot, "FFFFFF")
                        time_cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
                        time_cell.border = thin_border

                        # Числитель (верх)
                        for group_idx, group in enumerate(groups):
                            col = start_col + (group_idx * 3)

                            type_cell = ws.cell(row=time_row, column=col)
                            type_cell.border = thin_border
                            type_cell.fill = empty_type_fill
                            type_cell.alignment = Alignment(horizontal="center", vertical="center")
                            type_cell.font = Font(bold=True)

                            ws.cell(row=time_row, column=col + 1).border = thin_border
                            ws.cell(row=time_row, column=col + 2).border = thin_border

                            if group in transformed_data[course] and day in transformed_data[course][group]:
                                for slot_key, slot_data in transformed_data[course][group][day].items():
                                    clean_time = slot_data["Время"].replace(":", ".").replace(" - ", "-")
                                    if clean_time == time_slot and slot_data["Тип недели"] in ["Чс", "Чс/Зн"]:
                                        # Метка типа недели: "Чс/Зн" -> показываем "Чс" сверху
                                        week_type = slot_data["Тип недели"]
                                        type_label = "Чс" if week_type == "Чс/Зн" else week_type
                                        type_cell.value = type_label
                                        type_cell.fill = type_colors.get(week_type if week_type != "Чс/Зн" else "Чс")
                                        type_cell.font = type_font
                                        type_cell.alignment = Alignment(horizontal="center", vertical="center")

                                        # Предмет
                                        subject = slot_data["Предмет"]
                                        teacher = slot_data["Преподаватель"]
                                        lesson_type = slot_data.get("Тип пары", "")
                                        if lesson_type in ('лаба', 'семинар', 'лекция'):
                                            lesson_dict = {'лаба': 'ЛР', 'семинар': 'Семинар', 'лекция': 'Лекция'}
                                            lesson_type = lesson_dict[lesson_type]
                                        if teacher and lesson_type:
                                            cell_value = f"{subject}\n{teacher}\n{lesson_type}"
                                        elif teacher:
                                            cell_value = f"{subject}\n{teacher}"
                                        elif lesson_type:
                                            cell_value = f"{subject}\n{lesson_type}"
                                        else:
                                            cell_value = subject

                                        subject_cell = ws.cell(row=time_row, column=col + 1, value=cell_value)
                                        subject_cell.alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)

                                        classroom = slot_data.get("classroom", "-")
                                        ws.cell(row=time_row, column=col + 2, value=classroom).alignment = Alignment(horizontal="center")

                        # Знаменатель (низ)
                        for group_idx, group in enumerate(groups):
                            col = start_col + (group_idx * 3)

                            type_cell = ws.cell(row=time_row + 1, column=col)
                            type_cell.border = thin_border
                            type_cell.fill = empty_type_fill
                            type_cell.alignment = Alignment(horizontal="center", vertical="center")

                            ws.cell(row=time_row + 1, column=col + 1).border = thin_border
                            ws.cell(row=time_row + 1, column=col + 2).border = thin_border

                            if group in transformed_data[course] and day in transformed_data[course][group]:
                                for slot_key, slot_data in transformed_data[course][group][day].items():
                                    clean_time = slot_data["Время"].replace(":", ".").replace(" - ", "-")
                                    if clean_time == time_slot and slot_data["Тип недели"] in ["Зн", "Чс/Зн"]:
                                        # Метка типа недели: "Чс/Зн" -> показываем "Зн" снизу
                                        week_type = slot_data["Тип недели"]
                                        type_label = "Зн" if week_type == "Чс/Зн" else week_type
                                        type_cell.value = type_label
                                        type_cell.fill = type_colors.get(week_type if week_type != "Чс/Зн" else "Зн")
                                        type_cell.font = type_font
                                        type_cell.alignment = Alignment(horizontal="center", vertical="center")

                                        subject = slot_data["Предмет"]
                                        teacher = slot_data["Преподаватель"]
                                        lesson_type = slot_data.get("Тип пары", "")
                                        if lesson_type in ('лаба', 'семинар', 'лекция'):
                                            lesson_dict = {'лаба': 'ЛР', 'семинар': 'Семинар', 'лекция': 'Лекция'}
                                            lesson_type = lesson_dict[lesson_type]

                                        if teacher and lesson_type:
                                            cell_value = f"{subject}\n{teacher}\n{lesson_type}"
                                        elif teacher:
                                            cell_value = f"{subject}\n{teacher}"
                                        elif lesson_type:
                                            cell_value = f"{subject}\n{lesson_type}"
                                        else:
                                            cell_value = subject

                                        subject_cell = ws.cell(row=time_row + 1, column=col + 1, value=cell_value)
                                        subject_cell.alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)

                                        classroom = slot_data.get("classroom", "-")
                                        ws.cell(row=time_row + 1, column=col + 2, value=classroom).alignment = Alignment(horizontal="center")

                            # тонкий разделитель между "Чс" (верх) и "Зн" (низ) в колонке "Тип"
                            top_type_cell    = ws.cell(row=time_row,     column=col)
                            bottom_type_cell = ws.cell(row=time_row + 1, column=col)
                            with_border(top_type_cell,    b=thin_side)
                            with_border(bottom_type_cell, t=thin_side)

                        # Жирная линия-разделитель ПОСЛЕ каждой пары: от "Время" до правого края таблицы
                        last_col = start_col + len(groups) * 3 - 1  # последняя "Ауд."
                        for bcol in range(2, last_col + 1):         # начиная с колонки B ("Время")
                            cell = ws.cell(row=time_row + 1, column=bcol)
                            with_border(cell, b=thick_side)

                    current_row += 14

                    # Жирные границы между днями (как было)
                    for col in range(1, len(groups) * 3 + 3):
                        ws.cell(row=current_row, column=col).border = Border(top=Side(style="thick"))
                        ws.cell(row=current_row + 1, column=col).border = Border(bottom=Side(style="thick"))

                    current_row += 1  # Переход к следующему дню

                    # Толстые вертикальные границы между группами (как было)
                    for group_idx in range(len(groups)):
                        col = start_col + (group_idx * 3)
                        for row in range(1, current_row):
                            ws.cell(row=row, column=col).border = Border(left=Side(style="medium"))
                            ws.cell(row=row, column=col + 2).border = Border(right=Side(style="medium"))

                # Высота строк
                for row in range(1, current_row):
                    ws.row_dimensions[row].height = 70

        # Закрепляем верхнюю строку
        for course in range(1, 7):
            ws = wb[wb.sheetnames[course - 1]]
            ws.freeze_panes = "A2"

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        response = make_response(output.getvalue())
        response.headers["Content-Type"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        response.headers["Content-Disposition"] = "attachment; filename=schedule.xlsx"
        return response
    return []
